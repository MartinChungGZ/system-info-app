#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
系统信息采集工具 - System Information Collection Tool
跨平台桌面应用 (Windows / macOS)
免安装，PyInstaller打包后双击即可运行
"""

import tkinter as tk
from tkinter import ttk, messagebox, scrolledtext
import threading
import platform
import subprocess
import os
import sys
import re
import json
from datetime import datetime
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders
import ssl
import time
import urllib.request
import urllib.error

# ── Optional imports with fallback ──────────────────────────────────────────
try:
    import psutil
    HAS_PSUTIL = True
except ImportError:
    HAS_PSUTIL = False

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# Fix: PyInstaller --windowed sets sys.stdout/stderr to None, which crashes
# speedtest-cli's _Py3Utf8Output wrapper. Redirect to devnull as a safe fallback.
if sys.stdout is None:
    sys.stdout = open(os.devnull, 'w')
if sys.stderr is None:
    sys.stderr = open(os.devnull, 'w')

try:
    import speedtest as speedtest_lib
    HAS_SPEEDTEST = True
except ImportError:
    HAS_SPEEDTEST = False

# ── Constants ───────────────────────────────────────────────────────────────
APP_NAME = "系统信息采集工具"

# Determine app directory: works both in dev and PyInstaller bundled mode
if getattr(sys, 'frozen', False):
    # PyInstaller bundled: use the directory containing the executable
    APP_DIR = os.path.dirname(os.path.abspath(sys.executable))
else:
    APP_DIR = os.path.dirname(os.path.abspath(__file__))

CONFIG_FILE = os.path.join(APP_DIR, "config.json")

def _get_desktop_path():
    """Cross-platform desktop directory detection."""
    home = os.path.expanduser("~")
    for name in ["Desktop", "桌面"]:
        p = os.path.join(home, name)
        if os.path.isdir(p):
            return p
    return home

RECIPIENT_EMAIL = "zhongwenjian@zy.com"

# ── Helper: Log to UI ──────────────────────────────────────────────────────
_ui_log_func = None

def set_log_func(func):
    global _ui_log_func
    _ui_log_func = func

def log(msg):
    """Thread-safe logging to UI."""
    timestamp = datetime.now().strftime("%H:%M:%S")
    text = f"[{timestamp}] {msg}"
    if _ui_log_func:
        _ui_log_func(text)
    else:
        print(text)

# ── Config Management ──────────────────────────────────────────────────────
DEFAULT_CONFIG = {
    "smtp_server": "smtp.qq.com",
    "smtp_port": 465,
    "sender_email": "",
    "sender_password": "",
    "use_ssl": True
}

def load_config():
    if os.path.exists(CONFIG_FILE):
        try:
            with open(CONFIG_FILE, "r", encoding="utf-8") as f:
                cfg = json.load(f)
            for k, v in DEFAULT_CONFIG.items():
                if k not in cfg:
                    cfg[k] = v
            return cfg
        except Exception:
            return dict(DEFAULT_CONFIG)
    return dict(DEFAULT_CONFIG)

def save_config(cfg):
    try:
        with open(CONFIG_FILE, "w", encoding="utf-8") as f:
            json.dump(cfg, f, ensure_ascii=False, indent=2)
        return True
    except Exception as e:
        log(f"保存配置失败: {e}")
        return False

# ═══════════════════════════════════════════════════════════════════════════════
# 1. CPU Information Collection
# ═══════════════════════════════════════════════════════════════════════════════

def _get_friendly_cpu_name():
    """Get the friendly CPU brand/model name, cross-platform."""
    sys_type = platform.system()
    cpu_model = None

    try:
        if sys_type == "Windows":
            # Primary: Windows Registry (most reliable for friendly name)
            try:
                import winreg
                key = winreg.OpenKey(winreg.HKEY_LOCAL_MACHINE,
                    r"HARDWARE\DESCRIPTION\System\CentralProcessor\0")
                cpu_model, _ = winreg.QueryValueEx(key, "ProcessorNameString")
                winreg.CloseKey(key)
            except Exception:
                pass

            # Fallback: wmic (handle UTF-16 encoding)
            if not cpu_model:
                try:
                    result = subprocess.run(
                        ["wmic", "cpu", "get", "Name"],
                        capture_output=True, timeout=10
                    )
                    # wmic outputs UTF-16, decode manually
                    output = result.stdout.decode("utf-16-le", errors="ignore")
                    lines = output.strip().split("\n")
                    if len(lines) >= 2:
                        cpu_model = lines[1].strip()
                    if not cpu_model or "Name" in cpu_model:
                        cpu_model = None
                except Exception:
                    pass

        elif sys_type == "Darwin":
            result = subprocess.run(
                ["sysctl", "-n", "machdep.cpu.brand_string"],
                capture_output=True, text=True, timeout=10
            )
            cpu_model = result.stdout.strip()

        elif sys_type == "Linux":
            with open("/proc/cpuinfo", "r") as f:
                for line in f:
                    if "model name" in line:
                        cpu_model = line.split(":")[1].strip()
                        break
    except Exception:
        pass

    if not cpu_model or cpu_model == "Unknown":
        cpu_model = platform.processor() or "Unknown"

    return cpu_model

def _parse_cpu_brand(full_name):
    """Extract a clean brand name from full CPU identifier.
    e.g. '12th Gen Intel(R) Core(TM) i7-1260P' -> 'Intel Core i7-1260P'
         'AMD Ryzen 7 5800H with Radeon Graphics' -> 'AMD Ryzen 7 5800H'
    """
    if not full_name or full_name == "Unknown":
        return "Unknown"

    name = full_name.strip()
    # Remove registered/trademark symbols and extra whitespace
    name = re.sub(r'\s*\(R\)\s*', ' ', name)
    name = re.sub(r'\s*\(TM\)\s*', ' ', name)
    name = re.sub(r'\s+', ' ', name)

    # Try to extract Intel Core / AMD Ryzen pattern
    # Intel: match patterns like "Intel Core i7-1260P" or "Intel Core Ultra 7 155H"
    intel_match = re.search(
        r'(Intel\s*)?Core\s*(Ultra\s*)?(i[3579]|m[3579])[\s-]*\d+\w*',
        name, re.IGNORECASE
    )
    if intel_match:
        brand = intel_match.group(0).strip()
        if not brand.lower().startswith('intel'):
            brand = 'Intel ' + brand
        return brand

    # AMD: match "AMD Ryzen 7 5800H" style
    amd_match = re.search(
        r'AMD\s+Ryzen\s+(Threadripper\s+)?[3579]\s+\d+\w*',
        name, re.IGNORECASE
    )
    if amd_match:
        return amd_match.group(0).strip()

    # Apple Silicon
    apple_match = re.search(r'Apple\s+M\d+\s*(Pro|Max|Ultra)?', name, re.IGNORECASE)
    if apple_match:
        return apple_match.group(0).strip()

    # Fallback: remove "Gen" prefix, "with Radeon Graphics" suffix etc.
    name = re.sub(r'\d+th\s*Gen\s*', '', name)
    name = re.sub(r'\s*with\s+.*$', '', name)
    name = re.sub(r'\s*@\s*.*$', '', name)
    name = re.sub(r'\s+', ' ', name).strip()

    return name

def get_cpu_info():
    """Collect CPU brand, model, core count, and frequency."""
    info = {}
    info["操作系统"] = platform.system()
    info["系统版本"] = platform.version()
    info["系统架构"] = platform.machine()

    if HAS_PSUTIL:
        try:
            # CPU frequency
            freq = psutil.cpu_freq()
            if freq:
                info["CPU当前频率(MHz)"] = f"{freq.current:.1f}"
                info["CPU最大频率(MHz)"] = f"{freq.max:.1f}" if freq.max else "N/A"
            else:
                info["CPU当前频率(MHz)"] = "N/A"
                info["CPU最大频率(MHz)"] = "N/A"

            # Core count
            info["CPU物理核心数"] = str(psutil.cpu_count(logical=False))
            info["CPU逻辑核心数"] = str(psutil.cpu_count(logical=True))

            # CPU model - try to get friendly name from various sources
            cpu_model = _get_friendly_cpu_name()
            info["CPU型号"] = cpu_model
            # Also provide a short brand name like "Intel Core i7-1260P"
            info["CPU品牌"] = _parse_cpu_brand(cpu_model)
            return info
        except Exception as e:
            log(f"[WARN] psutil获取CPU信息失败: {e}")

    # ── Fallback: platform-specific commands ──
    return _get_cpu_info_fallback()

def _get_cpu_info_fallback():
    info = {}
    sys_type = platform.system()

    try:
        if sys_type == "Windows":
            cpu_model = _get_friendly_cpu_name()
            info["CPU型号"] = cpu_model
            info["CPU品牌"] = _parse_cpu_brand(cpu_model)
            # Get core counts via environment or other means
            info["CPU物理核心数"] = str(os.cpu_count() or "N/A")
            info["CPU逻辑核心数"] = str(os.cpu_count() or "N/A")

        elif sys_type == "Darwin":
            r = subprocess.run(["sysctl", "-n", "machdep.cpu.brand_string"],
                               capture_output=True, text=True, timeout=5)
            cpu_model = r.stdout.strip()
            info["CPU型号"] = cpu_model
            info["CPU品牌"] = _parse_cpu_brand(cpu_model)
            r2 = subprocess.run(["sysctl", "-n", "hw.physicalcpu"],
                                capture_output=True, text=True, timeout=5)
            info["CPU物理核心数"] = r2.stdout.strip()
            r3 = subprocess.run(["sysctl", "-n", "hw.logicalcpu"],
                                capture_output=True, text=True, timeout=5)
            info["CPU逻辑核心数"] = r3.stdout.strip()
            r4 = subprocess.run(["sysctl", "-n", "hw.cpufrequency"],
                                capture_output=True, text=True, timeout=5)
            try:
                hz = int(r4.stdout.strip())
                info["CPU当前频率(MHz)"] = f"{hz / 1_000_000:.0f}"
            except Exception:
                info["CPU当前频率(MHz)"] = "N/A"

        elif sys_type == "Linux":
            with open("/proc/cpuinfo", "r") as f:
                cpuinfo_text = f.read()
            for line in cpuinfo_text.split("\n"):
                if "model name" in line:
                    cpu_model = line.split(":")[1].strip()
                    info["CPU型号"] = cpu_model
                    info["CPU品牌"] = _parse_cpu_brand(cpu_model)
                elif "cpu MHz" in line:
                    info["CPU当前频率(MHz)"] = line.split(":")[1].strip()
                elif "cpu cores" in line:
                    info["CPU物理核心数"] = line.split(":")[1].strip()
            # Logical cores from nproc
            r = subprocess.run(["nproc"], capture_output=True, text=True, timeout=5)
            info["CPU逻辑核心数"] = r.stdout.strip()
    except Exception as e:
        log(f"[WARN] Fallback CPU detection failed: {e}")
        info["CPU型号"] = platform.processor() or "Unknown"

    # Fill defaults
    info.setdefault("CPU型号", "Unknown")
    info.setdefault("CPU品牌", "Unknown")
    info.setdefault("CPU物理核心数", "N/A")
    info.setdefault("CPU逻辑核心数", "N/A")
    info.setdefault("CPU当前频率(MHz)", "N/A")
    info.setdefault("CPU最大频率(MHz)", "N/A")
    return info

# ═══════════════════════════════════════════════════════════════════════════════
# 2. Memory Information Collection
# ═══════════════════════════════════════════════════════════════════════════════

def get_memory_info():
    """Collect total physical memory in GB."""
    info = {}

    if HAS_PSUTIL:
        try:
            mem = psutil.virtual_memory()
            total_bytes = mem.total
            info["物理内存总量(GB)"] = f"{total_bytes / (1024**3):.2f}"
            info["可用内存(GB)"] = f"{mem.available / (1024**3):.2f}"
            return info
        except Exception as e:
            log(f"[WARN] psutil获取内存失败: {e}")

    return _get_memory_info_fallback()

def _get_memory_info_fallback():
    info = {}
    sys_type = platform.system()

    try:
        if sys_type == "Windows":
            r = subprocess.run(
                ["wmic", "computersystem", "get", "TotalPhysicalMemory"],
                capture_output=True, text=True, timeout=10
            )
            lines = r.stdout.strip().split("\n")
            if len(lines) >= 2:
                val = lines[1].strip()
                info["物理内存总量(GB)"] = f"{int(val) / (1024**3):.2f}"

        elif sys_type == "Darwin":
            r = subprocess.run(["sysctl", "-n", "hw.memsize"],
                               capture_output=True, text=True, timeout=5)
            val = int(r.stdout.strip())
            info["物理内存总量(GB)"] = f"{val / (1024**3):.2f}"

        elif sys_type == "Linux":
            with open("/proc/meminfo", "r") as f:
                for line in f:
                    if "MemTotal" in line:
                        kb = int(re.findall(r"\d+", line)[0])
                        info["物理内存总量(GB)"] = f"{kb / (1024**2):.2f}"
                        break
    except Exception as e:
        log(f"[WARN] Fallback memory detection failed: {e}")

    info.setdefault("物理内存总量(GB)", "N/A")
    return info

# ═══════════════════════════════════════════════════════════════════════════════
# 3. Network Speed Test
# ═══════════════════════════════════════════════════════════════════════════════

def get_network_speed():
    """
    Test network download and upload speed with improved accuracy.
    Uses multi-threaded HTTP download + speedtest-cli for the most reliable result.
    """
    info = {}
    speedtest_ok = False

    # ── Method 1: speedtest-cli ──
    if HAS_SPEEDTEST:
        try:
            log("正在使用Speedtest进行网络测速...")
            st = speedtest_lib.Speedtest()

            # Filter for mainland China servers aggressively
            servers_dict = st.get_servers([])
            cn_servers = []
            all_servers = []
            for server_list in servers_dict.values():
                for s in server_list:
                    all_servers.append(s)
                    if s.get("cc") == "CN":
                        cn_servers.append(s)

            if cn_servers:
                log(f"从 {len(cn_servers)} 个中国测速服务器中选择最佳节点...")
                # Sort by latency, pick top 5, then let speedtest pick best
                cn_sorted = sorted(cn_servers, key=lambda s: s.get("latency", 99999))
                top_ids = [s["id"] for s in cn_sorted[:5]]
                st.get_servers(top_ids)
                st.get_best_server()
            else:
                log("未发现中国测速服务器，从所有服务器中选择...")
                st.get_best_server()

            server = st.results.server
            log(f"测速服务器: {server.get('sponsor', '?')} "
                f"[{server.get('cc', '?')}] 延迟={st.results.ping:.0f}ms")

            log("测试下载速率(Download)...")
            dl = st.download()
            info["下载速率(Mbps)"] = f"{dl / 1_000_000:.2f}"
            log(f"  下载: {dl / 1_000_000:.2f} Mbps")

            log("测试上传速率(Upload)...")
            ul = st.upload()
            info["上传速率(Mbps)"] = f"{ul / 1_000_000:.2f}"
            log(f"  上传: {ul / 1_000_000:.2f} Mbps")

            info["测速服务器"] = server.get("sponsor", "Unknown")
            info["延迟(ms)"] = f"{st.results.ping:.1f}"
            info["测速方式"] = f"Speedtest ({server.get('cc', 'Unknown')})"
            speedtest_ok = True
        except Exception as e:
            log(f"[WARN] Speedtest失败: {e}")

    # ── Method 2: Multi-threaded HTTP download (always run as validation) ──
    log("使用多线程HTTP下载进行验证测速...")
    http_info = _http_speed_test_multi()

    if not speedtest_ok:
        info.update(http_info)
    else:
        # Show HTTP results for comparison
        http_dl = http_info.get("下载速率(Mbps)", "N/A")
        log(f"多线程HTTP下载参考: {http_dl} Mbps")
        # If speedtest result seems off, prefer the better one
        try:
            if http_dl != "N/A" and http_dl != "测速失败":
                speedtest_dl = float(info.get("下载速率(Mbps)", 0))
                http_dl_val = float(http_dl)
                if http_dl_val > speedtest_dl * 2:
                    log(f"HTTP多线程测速({http_dl_val:.1f}Mbps)远高于Speedtest({speedtest_dl:.1f}Mbps)，"
                        f"采用多线程结果")
                    info["下载速率(Mbps)"] = f"{http_dl_val:.2f}"
                    info["测速方式"] = info.get("测速方式", "") + " + 多线程HTTP"
        except Exception:
            pass

    return info

def _http_speed_test_multi():
    """Multi-threaded HTTP speed test optimized for Chinese networks.

    Priority: domestic CDN endpoints → Cloudflare (global fallback).
    """
    from concurrent.futures import ThreadPoolExecutor, as_completed

    info = {}

    # ── Tier-1: Domestic Chinese CDN endpoints (fastest for mainland users) ──
    # These are large enough to measure real throughput and support Range requests
    domestic_configs = [
        # China Telecom Shanghai speed test
        ("http://speedtest1.online.sh.cn:8080/download/?size=25000000",
         "中国电信上海", False),
        # Tsinghua University mirror (10 MB test file)
        ("https://mirrors.tuna.tsinghua.edu.cn/ubuntu-releases/24.04/ubuntu-24.04-desktop-amd64.iso.zsync",
         "清华TUNA镜像站", False),
        # Alibaba Cloud OSS public test file
        ("https://alibaba.github.io/arthas/arthas-boot.jar",
         "阿里云OSS", False),
    ]

    # ── Tier-2: Cloudflare global CDN (supports Range, good multi-threading) ──
    cloudflare_configs = [
        ("https://speed.cloudflare.com/__down?bytes=25000000", "Cloudflare", True),
    ]

    best_dl = 0
    best_name = ""

    # Try domestic endpoints first, then Cloudflare
    all_configs = domestic_configs + cloudflare_configs

    for url, name, use_threads in all_configs:
        try:
            if use_threads:
                threads = 4
                log(f"  通过{name}多线程({threads}路)下载测试...")

                def download_chunk(url, timeout_val):
                    req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
                    ctx = ssl.create_default_context()
                    downloaded = 0
                    max_bytes = 8 * 1024 * 1024  # Cap per thread at 8MB
                    with urllib.request.urlopen(req, context=ctx, timeout=timeout_val) as resp:
                        while downloaded < max_bytes:
                            chunk = resp.read(65536)
                            if not chunk:
                                break
                            downloaded += len(chunk)
                    return downloaded

                start = time.time()
                total_bytes = 0
                with ThreadPoolExecutor(max_workers=threads) as executor:
                    futures = [executor.submit(download_chunk, url, 15) for _ in range(threads)]
                    for f in as_completed(futures):
                        try:
                            total_bytes += f.result(timeout=15)
                        except Exception:
                            pass

                elapsed = time.time() - start
                if elapsed > 0.5 and total_bytes > 500000:
                    dl = (total_bytes * 8) / elapsed / 1_000_000
                    log(f"  {name}多线程: {dl:.2f} Mbps ({total_bytes/(1024*1024):.0f}MB/{elapsed:.1f}s)")
                    if dl > best_dl:
                        best_dl = dl
                        best_name = name
                    break  # Cloudflare multi-thread is reliable, stop after first success
            else:
                # Single-threaded for domestic endpoints
                log(f"  通过{name}单线程下载测试...")
                dl_str = _single_thread_download_url(url, name)
                if dl_str != "测速失败":
                    dl_val = float(dl_str)
                    log(f"  {name}: {dl_val:.2f} Mbps")
                    if dl_val > best_dl:
                        best_dl = dl_val
                        best_name = name
                    # If we got a good result from domestic, don't bother with Cloudflare
                    if dl_val > 5:
                        break
        except Exception as e:
            log(f"  [WARN] {name}测速失败: {e}")
            continue

    if best_dl > 0:
        info["下载速率(Mbps)"] = f"{best_dl:.2f}"
        info["测速方式"] = f"HTTP测速 ({best_name})"
    else:
        # Fallback to single-threaded
        info["下载速率(Mbps)"] = _single_thread_download_fallback()
        info["测速方式"] = "HTTP测速(单线程)"

    # Upload test
    info["上传速率(Mbps)"] = _test_upload_speed()

    info["延迟(ms)"] = "N/A"
    return info

def _single_thread_download_url(url, name_hint=""):
    """Download from a single URL and return speed string."""
    try:
        req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
        ctx = ssl.create_default_context()
        start = time.time()
        downloaded = 0
        max_bytes = 15 * 1024 * 1024
        with urllib.request.urlopen(req, context=ctx, timeout=30) as resp:
            while downloaded < max_bytes:
                chunk = resp.read(65536)
                if not chunk:
                    break
                downloaded += len(chunk)
        elapsed = time.time() - start
        if elapsed > 0 and downloaded > 500000:
            return f"{(downloaded * 8) / elapsed / 1_000_000:.2f}"
    except Exception:
        pass
    return "测速失败"

def _single_thread_download_fallback():
    """Try multiple URLs for single-threaded download, domestic first."""
    test_urls = [
        # Domestic Chinese CDN endpoints (fastest for mainland users)
        ("http://speedtest1.online.sh.cn:8080/download/?size=25000000", "中国电信上海"),
        ("https://mirrors.tuna.tsinghua.edu.cn/ubuntu-releases/24.04/ubuntu-24.04-desktop-amd64.iso.zsync", "清华TUNA"),
        ("https://alibaba.github.io/arthas/arthas-boot.jar", "阿里云OSS"),
        # Global CDN fallbacks
        ("https://speed.cloudflare.com/__down?bytes=25000000", "Cloudflare"),
        ("https://proof.ovh.net/files/10Mb.dat", "OVH"),
    ]
    for url, name in test_urls:
        result = _single_thread_download_url(url, name)
        if result != "测速失败":
            return result
    return "测速失败"

def _test_upload_speed():
    """Test upload speed by posting data to test endpoints.

    Tries multiple upload endpoints for reliability, with domestic-friendly options.
    """
    test_endpoints = [
        # postman-echo (often has better connectivity in Asia-Pacific)
        "https://postman-echo.com/post",
        # httpbin (global, may be slower from China)
        "https://httpbin.org/post",
        # httpbin over plain HTTP (avoids TLS overhead for pure speed measurement)
        "http://httpbin.org/post",
    ]

    test_sizes = [1 * 1024 * 1024, 3 * 1024 * 1024]  # 1MB then 3MB
    best = 0

    for endpoint in test_endpoints:
        for size in test_sizes:
            try:
                test_data = b"0" * size
                req = urllib.request.Request(
                    endpoint,
                    data=test_data,
                    headers={
                        "Content-Type": "application/octet-stream",
                        "User-Agent": "Mozilla/5.0"
                    }
                )
                ctx = ssl.create_default_context() if endpoint.startswith("https") else None
                start = time.time()
                if ctx:
                    urllib.request.urlopen(req, context=ctx, timeout=20)
                else:
                    urllib.request.urlopen(req, timeout=20)
                elapsed = time.time() - start
                if elapsed > 0:
                    speed = (len(test_data) * 8) / elapsed / 1_000_000
                    if speed > best:
                        best = speed
            except Exception:
                continue

        # If we got a good result from this endpoint, no need to try others
        if best > 1:
            break

    if best > 0:
        return f"{best:.2f}"
    return "测速失败"

# ═══════════════════════════════════════════════════════════════════════════════
# 4 & 5. Browser Version Detection
# ═══════════════════════════════════════════════════════════════════════════════

def get_browser_versions():
    """Collect Google Chrome and Firefox versions."""
    chrome_version = _get_chrome_version()
    firefox_version = _get_firefox_version()
    return {
        "Chrome版本": chrome_version,
        "Firefox版本": firefox_version
    }

def _get_chrome_version():
    sys_type = platform.system()
    try:
        if sys_type == "Windows":
            import winreg

            # 1) Try Uninstall registry key (most reliable, works for per-user installs)
            for root, key_path in [
                (winreg.HKEY_CURRENT_USER,
                 r"Software\Microsoft\Windows\CurrentVersion\Uninstall\Google Chrome"),
                (winreg.HKEY_LOCAL_MACHINE,
                 r"Software\Microsoft\Windows\CurrentVersion\Uninstall\Google Chrome"),
                (winreg.HKEY_LOCAL_MACHINE,
                 r"Software\WOW6432Node\Microsoft\Windows\CurrentVersion\Uninstall\Google Chrome"),
            ]:
                try:
                    key = winreg.OpenKey(root, key_path)
                    val, _ = winreg.QueryValueEx(key, "DisplayVersion")
                    winreg.CloseKey(key)
                    if val:
                        return val
                except Exception:
                    pass

            # 2) Try BLBeacon registry
            for path in [
                r"SOFTWARE\Google\Chrome\BLBeacon",
                r"SOFTWARE\WOW6432Node\Google\Chrome\BLBeacon",
            ]:
                try:
                    key = winreg.OpenKey(winreg.HKEY_CURRENT_USER, path)
                    val, _ = winreg.QueryValueEx(key, "version")
                    winreg.CloseKey(key)
                    if val:
                        return val
                except Exception:
                    pass

            # 3) Try known install paths
            chrome_paths = [
                r"C:\Program Files\Google\Chrome\Application\chrome.exe",
                r"C:\Program Files (x86)\Google\Chrome\Application\chrome.exe",
            ]
            local_appdata = os.environ.get("LOCALAPPDATA", "")
            if local_appdata:
                chrome_paths.insert(0, os.path.join(local_appdata,
                    r"Google\Chrome\Application\chrome.exe"))

            for exe in chrome_paths:
                if not os.path.exists(exe):
                    continue
                # Try PowerShell to get file version (works even if Chrome is running)
                try:
                    ps_cmd = (
                        f'(Get-Item "{exe}").VersionInfo.ProductVersion'
                    )
                    r = subprocess.run(
                        ['powershell', '-NoProfile', '-Command', ps_cmd],
                        capture_output=True, text=True, timeout=15
                    )
                    ver = r.stdout.strip()
                    if ver and re.match(r'\d+\.\d+', ver):
                        return ver
                except Exception:
                    pass
                # Fallback: run chrome --version
                try:
                    r = subprocess.run([exe, "--version"], capture_output=True,
                                       text=True, timeout=10)
                    m = re.search(r"(\d+\.\d+\.\d+\.\d+)", r.stdout + r.stderr)
                    if m:
                        return m.group(1)
                except Exception:
                    pass

        elif sys_type == "Darwin":
            exe = "/Applications/Google Chrome.app/Contents/MacOS/Google Chrome"
            if os.path.exists(exe):
                r = subprocess.run([exe, "--version"], capture_output=True,
                                   text=True, timeout=10)
                m = re.search(r"(\d+\.\d+\.\d+\.\d+)", r.stdout + r.stderr)
                if m:
                    return m.group(1)
            # Try reading plist
            plist = "/Applications/Google Chrome.app/Contents/Info.plist"
            if os.path.exists(plist):
                r = subprocess.run(
                    ["/usr/libexec/PlistBuddy", "-c",
                     "Print :CFBundleShortVersionString", plist],
                    capture_output=True, text=True, timeout=5
                )
                if r.stdout.strip():
                    return r.stdout.strip()

        elif sys_type == "Linux":
            for cmd in ["google-chrome", "google-chrome-stable", "chromium",
                        "chromium-browser"]:
                try:
                    r = subprocess.run([cmd, "--version"], capture_output=True,
                                       text=True, timeout=5)
                    m = re.search(r"(\d+\.\d+\.\d+\.\d+)", r.stdout + r.stderr)
                    if m:
                        return m.group(1)
                except Exception:
                    continue
    except Exception as e:
        log(f"[WARN] Chrome版本检测失败: {e}")

    return "未安装或无法检测"

def _get_firefox_version():
    sys_type = platform.system()
    try:
        if sys_type == "Windows":
            # Try registry
            try:
                import winreg
                key = winreg.OpenKey(
                    winreg.HKEY_LOCAL_MACHINE,
                    r"SOFTWARE\Mozilla\Mozilla Firefox"
                )
                val, _ = winreg.QueryValueEx(key, "CurrentVersion")
                winreg.CloseKey(key)
                if val:
                    return val
            except Exception:
                pass

            # Try wmic
            r = subprocess.run(
                ['wmic', 'datafile', 'where',
                 'name="C:\\\\Program Files\\\\Mozilla Firefox\\\\firefox.exe"',
                 'get', 'Version'],
                capture_output=True, text=True, timeout=10
            )
            lines = r.stdout.strip().split("\n")
            if len(lines) >= 2 and lines[1].strip():
                return lines[1].strip()

            r = subprocess.run(
                ['wmic', 'datafile', 'where',
                 'name="C:\\\\Program Files (x86)\\\\Mozilla Firefox\\\\firefox.exe"',
                 'get', 'Version'],
                capture_output=True, text=True, timeout=10
            )
            lines = r.stdout.strip().split("\n")
            if len(lines) >= 2 and lines[1].strip():
                return lines[1].strip()

            for exe in [
                r"C:\Program Files\Mozilla Firefox\firefox.exe",
                r"C:\Program Files (x86)\Mozilla Firefox\firefox.exe",
            ]:
                if os.path.exists(exe):
                    r = subprocess.run([exe, "--version"], capture_output=True,
                                       text=True, timeout=10)
                    m = re.search(r"(\d+\.\d+(\.\d+)?)", r.stdout + r.stderr)
                    if m:
                        return m.group(1)

        elif sys_type == "Darwin":
            exe = "/Applications/Firefox.app/Contents/MacOS/firefox"
            if os.path.exists(exe):
                r = subprocess.run([exe, "--version"], capture_output=True,
                                   text=True, timeout=10)
                m = re.search(r"(\d+\.\d+(\.\d+)?)", r.stdout + r.stderr)
                if m:
                    return m.group(1)
            plist = "/Applications/Firefox.app/Contents/Info.plist"
            if os.path.exists(plist):
                r = subprocess.run(
                    ["/usr/libexec/PlistBuddy", "-c",
                     "Print :CFBundleShortVersionString", plist],
                    capture_output=True, text=True, timeout=5
                )
                if r.stdout.strip():
                    return r.stdout.strip()

        elif sys_type == "Linux":
            for cmd in ["firefox"]:
                try:
                    r = subprocess.run([cmd, "--version"], capture_output=True,
                                       text=True, timeout=5)
                    m = re.search(r"(\d+\.\d+(\.\d+)?)", r.stdout + r.stderr)
                    if m:
                        return m.group(1)
                except Exception:
                    continue
    except Exception as e:
        log(f"[WARN] Firefox版本检测失败: {e}")

    return "未安装或无法检测"

# ═══════════════════════════════════════════════════════════════════════════════
# 6. Excel File Generation
# ═══════════════════════════════════════════════════════════════════════════════

def create_excel(campus, name, cpu_info, mem_info, net_info, browser_info):
    """Create a simple flat-table Excel file for easy data processing."""
    if not HAS_OPENPYXL:
        raise RuntimeError("openpyxl未安装，无法生成Excel文件。请运行: pip install openpyxl")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "系统信息采集报告"

    # Simple styles
    header_font = Font(name="微软雅黑", size=11, bold=True)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    cell_font = Font(name="微软雅黑", size=10)
    border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")

    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 50

    # ── Flat data rows ──
    rows = [
        ("校区", campus),
        ("姓名", name),
        ("采集时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("操作系统", f"{platform.system()} {platform.release()} ({platform.machine()})"),
        ("主机名", platform.node()),
        ("CPU品牌", cpu_info.get("CPU品牌", "N/A")),
        ("CPU型号", cpu_info.get("CPU型号", "N/A")),
        ("CPU物理核心数", cpu_info.get("CPU物理核心数", "N/A")),
        ("CPU逻辑核心数", cpu_info.get("CPU逻辑核心数", "N/A")),
        ("CPU当前频率(MHz)", cpu_info.get("CPU当前频率(MHz)", "N/A")),
        ("CPU最大频率(MHz)", cpu_info.get("CPU最大频率(MHz)", "N/A")),
        ("物理内存总量(GB)", mem_info.get("物理内存总量(GB)", "N/A")),
        ("可用内存(GB)", mem_info.get("可用内存(GB)", "N/A")),
        ("Google Chrome版本", browser_info.get("Chrome版本", "N/A")),
        ("Mozilla Firefox版本", browser_info.get("Firefox版本", "N/A")),
        ("下载速率(Mbps)", net_info.get("下载速率(Mbps)", "N/A")),
        ("上传速率(Mbps)", net_info.get("上传速率(Mbps)", "N/A")),
        ("延迟(ms)", net_info.get("延迟(ms)", "N/A")),
        ("测速方式", net_info.get("测速方式", "N/A")),
    ]

    # Write header row
    for col, header in enumerate(["项目", "值"], 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center_align

    # Write data rows
    for i, (key, val) in enumerate(rows):
        r = i + 2
        cell_a = ws.cell(row=r, column=1, value=key)
        cell_a.font = cell_font
        cell_a.border = border
        cell_a.alignment = left_align

        cell_b = ws.cell(row=r, column=2, value=str(val))
        cell_b.font = cell_font
        cell_b.border = border
        cell_b.alignment = left_align

    # Generate filename: 校区-姓名.xlsx, sanitize for cross-platform safety
    safe_campus = re.sub(r'[\\/:*?"<>|]', '_', campus)
    safe_name = re.sub(r'[\\/:*?"<>|]', '_', name)
    filename = f"{safe_campus}-{safe_name}.xlsx"
    filepath = os.path.join(_get_desktop_path(), filename)

    try:
        wb.save(filepath)
        log(f"Excel文件已保存: {filepath}")
        return filepath
    except Exception as e:
        raise RuntimeError(f"保存Excel文件失败: {e}")

# ═══════════════════════════════════════════════════════════════════════════════
# 7. Email Sending
# ═══════════════════════════════════════════════════════════════════════════════

def send_email(config, campus, name, excel_path):
    """Send the Excel report via email."""
    smtp_server = config.get("smtp_server", "")
    smtp_port = int(config.get("smtp_port", 465))
    sender_email = config.get("sender_email", "")
    sender_password = config.get("sender_password", "")
    use_ssl = config.get("use_ssl", True)

    if not all([smtp_server, sender_email, sender_password]):
        raise ValueError("邮件配置不完整，请先设置SMTP服务器、发件邮箱和密码/授权码")

    subject = f"系统信息采集报告 - {campus} - {name} - {datetime.now().strftime('%Y%m%d_%H%M%S')}"
    body = f"""
    系统信息采集报告

    校区: {campus}
    姓名: {name}
    采集时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
    主机名: {platform.node()}

    详细信息请查看附件Excel文件。

    此邮件由系统信息采集工具自动发送。
    """

    msg = MIMEMultipart()
    msg["From"] = sender_email
    msg["To"] = RECIPIENT_EMAIL
    msg["Subject"] = subject
    msg.attach(MIMEText(body, "plain", "utf-8"))

    # Attach Excel file
    with open(excel_path, "rb") as f:
        part = MIMEBase("application", "vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        part.set_payload(f.read())
        encoders.encode_base64(part)
        part.add_header(
            "Content-Disposition",
            f'attachment; filename="{os.path.basename(excel_path)}"'
        )
        msg.attach(part)

    log(f"正在连接SMTP服务器 {smtp_server}:{smtp_port} ...")
    try:
        if use_ssl:
            context = ssl.create_default_context()
            with smtplib.SMTP_SSL(smtp_server, smtp_port, context=context, timeout=30) as server:
                server.login(sender_email, sender_password)
                server.sendmail(sender_email, [RECIPIENT_EMAIL], msg.as_string())
        else:
            with smtplib.SMTP(smtp_server, smtp_port, timeout=30) as server:
                server.starttls(context=ssl.create_default_context())
                server.login(sender_email, sender_password)
                server.sendmail(sender_email, [RECIPIENT_EMAIL], msg.as_string())

        log(f"邮件已成功发送到 {RECIPIENT_EMAIL}")
        return True
    except smtplib.SMTPAuthenticationError:
        raise RuntimeError("邮箱认证失败，请检查发件邮箱和密码/授权码是否正确")
    except smtplib.SMTPConnectError:
        raise RuntimeError(f"无法连接到SMTP服务器 {smtp_server}:{smtp_port}，请检查服务器地址和端口")
    except Exception as e:
        raise RuntimeError(f"邮件发送失败: {e}")

# ═══════════════════════════════════════════════════════════════════════════════
# GUI - Email Configuration Dialog
# ═══════════════════════════════════════════════════════════════════════════════

class EmailConfigDialog(tk.Toplevel):
    """Email configuration dialog."""

    SMTP_PRESETS = {
        "QQ邮箱": {"server": "smtp.qq.com", "port": 465, "ssl": True},
        "163邮箱": {"server": "smtp.163.com", "port": 465, "ssl": True},
        "126邮箱": {"server": "smtp.126.com", "port": 465, "ssl": True},
        "腾讯企业邮箱": {"server": "smtp.exmail.qq.com", "port": 465, "ssl": True},
        "网易企业邮箱": {"server": "smtp.ym.163.com", "port": 465, "ssl": True},
        "Gmail": {"server": "smtp.gmail.com", "port": 587, "ssl": False},
        "自定义": {"server": "", "port": 465, "ssl": True},
    }

    def __init__(self, parent, config):
        super().__init__(parent)
        self.title("邮箱设置")
        self.config = config
        self.result = None

        # Make modal
        self.transient(parent)
        self.grab_set()

        self._create_widgets()
        self._load_config()

        # Center on parent
        self.geometry("+%d+%d" % (
            parent.winfo_rootx() + 50,
            parent.winfo_rooty() + 50
        ))
        self.resizable(False, False)

    def _create_widgets(self):
        frame = ttk.Frame(self, padding=15)
        frame.pack(fill="both", expand=True)

        # Preset selection
        ttk.Label(frame, text="邮箱预设:").grid(row=0, column=0, sticky="w", pady=3)
        self.preset_var = tk.StringVar(value="QQ邮箱")
        self.preset_combo = ttk.Combobox(frame, textvariable=self.preset_var,
                                          values=list(self.SMTP_PRESETS.keys()),
                                          state="readonly", width=18)
        self.preset_combo.grid(row=0, column=1, sticky="ew", pady=3)
        self.preset_combo.bind("<<ComboboxSelected>>", self._on_preset_change)

        # SMTP Server
        ttk.Label(frame, text="SMTP服务器:").grid(row=1, column=0, sticky="w", pady=3)
        self.smtp_server = ttk.Entry(frame, width=30)
        self.smtp_server.grid(row=1, column=1, sticky="ew", pady=3)

        # SMTP Port
        ttk.Label(frame, text="SMTP端口:").grid(row=2, column=0, sticky="w", pady=3)
        self.smtp_port = ttk.Entry(frame, width=30)
        self.smtp_port.grid(row=2, column=1, sticky="ew", pady=3)

        # SSL/TLS
        self.ssl_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(frame, text="使用SSL加密连接",
                        variable=self.ssl_var).grid(row=3, column=0, columnspan=2,
                                                     sticky="w", pady=3)

        # Sender Email
        ttk.Label(frame, text="发件邮箱:").grid(row=4, column=0, sticky="w", pady=3)
        self.sender_email = ttk.Entry(frame, width=30)
        self.sender_email.grid(row=4, column=1, sticky="ew", pady=3)

        # Sender Password
        ttk.Label(frame, text="密码/授权码:").grid(row=5, column=0, sticky="w", pady=3)
        self.sender_password = ttk.Entry(frame, width=30, show="*")
        self.sender_password.grid(row=5, column=1, sticky="ew", pady=3)

        # Note
        note = ("说明: 使用QQ邮箱请填写授权码（在QQ邮箱设置→账户→POP3/SMTP服务中获取），"
                "163/126邮箱同样需要使用授权码。")
        note_lbl = ttk.Label(frame, text=note, wraplength=360, foreground="gray")
        note_lbl.grid(row=6, column=0, columnspan=2, sticky="w", pady=8)

        # Recipient (display only)
        ttk.Label(frame, text=f"收件邮箱: {RECIPIENT_EMAIL}",
                  foreground="blue").grid(row=7, column=0, columnspan=2,
                                           sticky="w", pady=3)

        # Buttons
        btn_frame = ttk.Frame(frame)
        btn_frame.grid(row=8, column=0, columnspan=2, pady=10)
        ttk.Button(btn_frame, text="保存", command=self._on_save).pack(side="left", padx=5)
        ttk.Button(btn_frame, text="取消", command=self.destroy).pack(side="left", padx=5)

    def _on_preset_change(self, event=None):
        preset = self.preset_var.get()
        p = self.SMTP_PRESETS.get(preset, {})
        self.smtp_server.delete(0, "end")
        self.smtp_server.insert(0, p.get("server", ""))
        self.smtp_port.delete(0, "end")
        self.smtp_port.insert(0, str(p.get("port", 465)))
        self.ssl_var.set(p.get("ssl", True))

    def _load_config(self):
        self.smtp_server.insert(0, self.config.get("smtp_server", ""))
        self.smtp_port.insert(0, str(self.config.get("smtp_port", 465)))
        self.ssl_var.set(self.config.get("use_ssl", True))
        self.sender_email.insert(0, self.config.get("sender_email", ""))
        self.sender_password.insert(0, self.config.get("sender_password", ""))

    def _on_save(self):
        cfg = {
            "smtp_server": self.smtp_server.get().strip(),
            "smtp_port": self.smtp_port.get().strip(),
            "sender_email": self.sender_email.get().strip(),
            "sender_password": self.sender_password.get().strip(),
            "use_ssl": self.ssl_var.get(),
        }

        if not cfg["sender_email"]:
            messagebox.showwarning("提示", "请填写发件邮箱", parent=self)
            return
        if not cfg["sender_password"]:
            messagebox.showwarning("提示", "请填写密码/授权码", parent=self)
            return
        if not cfg["smtp_server"]:
            messagebox.showwarning("提示", "请填写SMTP服务器", parent=self)
            return

        try:
            port = int(cfg["smtp_port"])
            cfg["smtp_port"] = port
        except ValueError:
            messagebox.showwarning("提示", "SMTP端口必须是数字", parent=self)
            return

        self.result = cfg
        self.destroy()

# ═══════════════════════════════════════════════════════════════════════════════
# GUI - Main Application Window
# ═══════════════════════════════════════════════════════════════════════════════

class MainApp(tk.Tk):
    def __init__(self):
        super().__init__()

        self.title(APP_NAME)
        self.geometry("680x580")
        self.resizable(True, True)
        self.minsize(580, 480)

        # Center window
        self.update_idletasks()
        sw = self.winfo_screenwidth()
        sh = self.winfo_screenheight()
        x = (sw - 680) // 2
        y = (sh - 580) // 2
        self.geometry(f"+{x}+{y}")

        # Icons
        try:
            if platform.system() == "Windows":
                self.iconbitmap(default="")
        except Exception:
            pass

        self.config = load_config()
        self._running = False

        self._create_widgets()

        # Set up logging
        set_log_func(self._log_to_ui)

        self.protocol("WM_DELETE_WINDOW", self._on_close)

    def _create_widgets(self):
        # ── Title ──
        title_frame = ttk.Frame(self)
        title_frame.pack(pady=(15, 5))

        title_lbl = tk.Label(title_frame, text=APP_NAME,
                             font=("微软雅黑", 18, "bold"),
                             fg="#1F4E79")
        title_lbl.pack()

        subtitle = tk.Label(title_frame, text="跨平台系统信息采集 | Windows & macOS",
                            font=("微软雅黑", 9), fg="gray")
        subtitle.pack()

        # ── Input Area ──
        input_frame = ttk.LabelFrame(self, text="信息填写", padding=15)
        input_frame.pack(fill="x", padx=20, pady=10)

        # Campus dropdown
        ttk.Label(input_frame, text="所属校区:").grid(row=0, column=0,
                                                        sticky="w", padx=(0, 10))
        self.campus_var = tk.StringVar()
        self.campus_combo = ttk.Combobox(
            input_frame,
            textvariable=self.campus_var,
            values=[
                "初四越秀校区", "初四梅花园校区", "初四海珠校区", "初四天河校区",
                "高四公园前校区", "高四海珠校区", "高四梅花园校区", "高四花都校区",
                "中高复中台校区"
            ],
            state="readonly",
            width=25,
            font=("微软雅黑", 11)
        )
        self.campus_combo.grid(row=0, column=1, sticky="w", pady=5)
        self.campus_combo.current(0)

        # Name input (max 100 chars, Chinese/English supported)
        ttk.Label(input_frame, text="姓名:").grid(row=1, column=0,
                                                   sticky="w", padx=(0, 10))
        self.name_var = tk.StringVar()
        vcmd = (self.register(self._validate_name), '%P')
        self.name_entry = ttk.Entry(input_frame, textvariable=self.name_var,
                                    width=28, font=("微软雅黑", 11),
                                    validate='key', validatecommand=vcmd)
        self.name_entry.grid(row=1, column=1, sticky="w", pady=5)

        # ── Buttons ──
        btn_frame = ttk.Frame(self)
        btn_frame.pack(fill="x", padx=20, pady=5)

        self.start_btn = ttk.Button(btn_frame, text="开始搜集",
                                    command=self._start_collection)
        self.start_btn.pack(side="left", padx=5)

        self.email_config_btn = ttk.Button(btn_frame, text="邮箱设置(暂不可用)",
                                           command=self._open_email_config,
                                           state="disabled")
        self.email_config_btn.pack(side="left", padx=5)

        # Status indicator
        self.status_var = tk.StringVar(value="就绪")
        status_lbl = ttk.Label(btn_frame, textvariable=self.status_var,
                               foreground="green")
        status_lbl.pack(side="right", padx=10)

        # ── Progress Bar ──
        self.progress = ttk.Progressbar(self, mode="determinate", length=600)
        self.progress.pack(fill="x", padx=20, pady=5)

        # ── Log Output ──
        log_frame = ttk.LabelFrame(self, text="运行日志", padding=5)
        log_frame.pack(fill="both", expand=True, padx=20, pady=10)

        self.log_text = scrolledtext.ScrolledText(
            log_frame, height=14, font=("Consolas", 9),
            wrap=tk.WORD, state="disabled", bg="#1e1e1e", fg="#d4d4d4"
        )
        self.log_text.pack(fill="both", expand=True)

        # Configure tags for colored text
        self.log_text.tag_config("success", foreground="#4ec94e")
        self.log_text.tag_config("error", foreground="#f44747")
        self.log_text.tag_config("warn", foreground="#e5c054")
        self.log_text.tag_config("info", foreground="#569cd6")

        # ── Version info ──
        ver_lbl = tk.Label(self, text="v1.0 | 支持 Windows / macOS",
                           font=("微软雅黑", 8), fg="gray")
        ver_lbl.pack(pady=3)

    def _log_to_ui(self, msg):
        """Thread-safe log to the scrolled text widget."""
        self.log_text.configure(state="normal")
        if "成功" in msg or "完成" in msg:
            self.log_text.insert("end", msg + "\n", "success")
        elif "失败" in msg or "错误" in msg:
            self.log_text.insert("end", msg + "\n", "error")
        elif "WARN" in msg:
            self.log_text.insert("end", msg + "\n", "warn")
        else:
            self.log_text.insert("end", msg + "\n")
        self.log_text.see("end")
        self.log_text.configure(state="disabled")

    def _open_email_config(self):
        EmailConfigDialog(self, self.config)
        self.config = load_config()  # Reload in case config was saved

    def _set_status(self, msg, color="black"):
        self.status_var.set(msg)
        if "成功" in msg or "完成" in msg:
            color = "green"
        elif "失败" in msg:
            color = "red"
        # tkinter StringVar can't have color directly - use label config
        for child in self.children.values():
            if isinstance(child, ttk.Frame):
                for gc in child.children.values():
                    if isinstance(gc, ttk.Label) and gc.cget("textvariable") == str(self.status_var):
                        gc.configure(foreground=color)

    def _validate_name(self, new_text):
        """Validate name input: max 100 characters."""
        return len(new_text) <= 100

    def _start_collection(self):
        campus = self.campus_var.get()
        name = self.name_var.get().strip()

        if not name:
            messagebox.showwarning("提示", "请输入姓名")
            return
        if not campus:
            messagebox.showwarning("提示", "请选择校区")
            return

        self._running = True
        self.start_btn.configure(state="disabled")
        self.progress["value"] = 0
        self._set_status("正在采集...")

        # Run in background thread
        cfg = load_config()
        thread = threading.Thread(target=self._collection_worker,
                                  args=(campus, name, cfg), daemon=True)
        thread.start()

    def _reset_buttons(self):
        self.start_btn.configure(state="normal")

    def _collection_worker(self, campus, name, cfg):
        """Main collection worker - runs in background thread."""
        steps = [
            (1, "CPU信息", self._collect_cpu),
            (2, "内存信息", self._collect_memory),
            (3, "网络测速", self._collect_network),
            (4, "浏览器版本(Chrome)", self._collect_chrome),
            (5, "浏览器版本(Firefox)", self._collect_firefox),
            (6, "生成Excel", self._collect_excel),
        ]

        results = {}
        failed = False

        for step_num, step_name, step_func in steps:
            if not self._running:
                log("采集已中止")
                break
            try:
                log(f"[{step_num}/6] 正在采集: {step_name}...")
                self.progress["value"] = step_num * (100 / 6)
                result = step_func(campus, name, cfg, results)
                if result:
                    results.update(result)
                log(f"[{step_num}/6] {step_name} - 完成 ✓")
                self._set_status(f"已完成 {step_num}/6: {step_name}")
            except Exception as e:
                log(f"[{step_num}/6] {step_name} - 失败 ✗: {e}")
                self._set_status(f"步骤{step_num}失败: {step_name}")
                failed = True
                if step_num == 6:
                    messagebox.showerror("错误", f"生成Excel文件失败:\n{e}")
                    break

        self.progress["value"] = 100

        if not failed:
            log("✓ 所有采集步骤完成！")
            self._set_status("采集完成")

        self.after(0, self._reset_buttons)
        self._running = False

    # ── Individual collection step functions ──

    def _collect_cpu(self, campus, name, cfg, results):
        log("检测CPU信息...")
        return {"cpu": get_cpu_info()}

    def _collect_memory(self, campus, name, cfg, results):
        log("检测内存信息...")
        return {"memory": get_memory_info()}

    def _collect_network(self, campus, name, cfg, results):
        log("正在进行网络测速（连接国内测速网络）...")
        return {"network": get_network_speed()}

    def _collect_chrome(self, campus, name, cfg, results):
        log("检测Chrome浏览器版本...")
        return {"chrome": _get_chrome_version()}

    def _collect_firefox(self, campus, name, cfg, results):
        log("检测Firefox浏览器版本...")
        return {"firefox": _get_firefox_version()}

    def _collect_excel(self, campus, name, cfg, results):
        log("生成Excel报告...")
        cpu_info = results.get("cpu", {})
        mem_info = results.get("memory", {})
        net_info = results.get("network", {})
        browser_info = {
            "Chrome版本": results.get("chrome", "N/A"),
            "Firefox版本": results.get("firefox", "N/A"),
        }
        excel_path = create_excel(campus, name, cpu_info, mem_info, net_info, browser_info)
        return {"excel_path": excel_path}

    def _collect_email(self, campus, name, cfg, results):
        excel_path = results.get("excel_path")
        if not excel_path:
            log("[INFO] Excel文件未生成，跳过发送")
            return {}
        if not cfg.get("sender_email") or not cfg.get("sender_password"):
            log("[INFO] 邮件未配置，跳过发送")
            return {}
        log("发送邮件...")
        send_email(cfg, campus, name, excel_path)
        return {}

    def _on_close(self):
        if self._running:
            if messagebox.askyesno("确认", "采集正在进行中，确定要退出吗？"):
                self._running = False
                self.destroy()
        else:
            self.destroy()

# ═══════════════════════════════════════════════════════════════════════════════
# Entry Point
# ═══════════════════════════════════════════════════════════════════════════════

def main():
    # Check for required non-optional deps
    missing = []
    if not HAS_OPENPYXL:
        missing.append("openpyxl")
    if not HAS_PSUTIL:
        missing.append("psutil")

    if missing:
        # Try to show GUI warning
        root = tk.Tk()
        root.withdraw()
        messagebox.showwarning(
            "缺少依赖",
            f"缺少以下Python包：{', '.join(missing)}\n\n"
            f"请运行以下命令安装：\n"
            f"pip install {' '.join(missing)}\n\n"
            f"某些功能将以降级模式运行。"
        )
        root.destroy()

    app = MainApp()
    app.mainloop()

if __name__ == "__main__":
    main()
